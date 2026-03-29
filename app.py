# -*- coding: utf-8 -*-
"""
Buscador de Bibliografía Científica Avanzado (VERSIÓN SERVIDOR WEB)
Con API Key de Groq, Soporte Bilingüe, Créditos y DOI.
"""

import os
import re
import requests
import urllib3
from io import BytesIO
from flask import Flask, request, send_file, render_template_string
import openpyxl
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup
import PyPDF2

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

TRANSLATIONS = {
    'es': {
        'title': 'Buscador de Bibliografía Científica',
        'lang_btn': 'English Version',
        'lang_url': '/?lang=en',
        'hero_title': '🔬 Extractor Científico Avanzado',
        'hero_sub': 'Búsqueda IA + Scraping de PDFs y Webs',
        'tut_title': '🔑 ¿Cómo obtener tu API Key gratuita?',
        'tut_p': 'Para que la Inteligencia Artificial genere las palabras clave científicas, necesitas una clave gratuita de Groq:',
        'tut_li1': 'Regístrate gratis en <a href="https://console.groq.com/" target="_blank">console.groq.com</a>.',
        'tut_li2': 'En el menú lateral, ve a <b>"API Keys"</b> y haz clic en <i>"Create API Key"</i>.',
        'tut_li3': 'Copia la clave generada (empieza con <code>gsk_</code>) y pégala en el recuadro de abajo.',
        'lbl_api': 'Tu API Key de Groq:',
        'ph_api': 'gsk_.......................................',
        'lbl_tema': 'Tema de investigación:',
        'ph_tema': 'Ej: Inteligencia Artificial en Medicina',
        'lbl_idioma': 'Idioma de los artículos:',
        'opt_es': 'Español',
        'opt_en': 'Inglés',
        'lbl_limite': 'Cantidad de artículos:',
        'btn_submit': 'Generar y Descargar Excel 📊',
        'loading_title': 'Extrayendo información...',
        'loading_text': 'Leyendo webs y PDFs. Esto puede tardar varios minutos.<br><b>No cierres ni recargues esta página.</b>',
        'cred_title': '👨‍🔬 Créditos y Citación',
        'cred_author': 'Elaborado por el <b>Dr. Utiel Zagada Dominguez</b>.',
        'cred_cite': 'Si esta herramienta le fue de utilidad, favor de citarlo como proveedor de herramienta tecnológica para buscar referencias bibliográficas.',
        'cred_doi': '<b>DOI de esta herramienta:</b> <a href="https://doi.org/10.XXXX/tu-doi-aqui" target="_blank">10.XXXX/tu-doi-aqui</a> <i>(Pendiente de asignar)</i>'
    },
    'en': {
        'title': 'Scientific Bibliography Search Engine',
        'lang_btn': 'Versión en Español',
        'lang_url': '/?lang=es',
        'hero_title': '🔬 Advanced Scientific Extractor',
        'hero_sub': 'AI Search + PDF & Web Scraping',
        'tut_title': '🔑 How to get your free API Key?',
        'tut_p': 'For the Artificial Intelligence to generate scientific keywords, you need a free Groq key:',
        'tut_li1': 'Sign up for free at <a href="https://console.groq.com/" target="_blank">console.groq.com</a>.',
        'tut_li2': 'In the sidebar, go to <b>"API Keys"</b> and click on <i>"Create API Key"</i>.',
        'tut_li3': 'Copy the generated key (starts with <code>gsk_</code>) and paste it in the box below.',
        'lbl_api': 'Your Groq API Key:',
        'ph_api': 'gsk_.......................................',
        'lbl_tema': 'Research topic:',
        'ph_tema': 'Ex: Artificial Intelligence in Medicine',
        'lbl_idioma': 'Language of articles:',
        'opt_es': 'Spanish',
        'opt_en': 'English',
        'lbl_limite': 'Amount of articles:',
        'btn_submit': 'Generate and Download Excel 📊',
        'loading_title': 'Extracting information...',
        'loading_text': 'Reading webs and PDFs. This might take several minutes.<br><b>Do not close or reload this page.</b>',
        'cred_title': '👨‍🔬 Credits & Citation',
        'cred_author': 'Developed by <b>Dr. Utiel Zagada Dominguez</b>.',
        'cred_cite': 'If you found this tool useful, please cite him as the provider of the technological tool to search for bibliographic references.',
        'cred_doi': '<b>DOI of this tool:</b> <a href="https://doi.org/10.XXXX/tu-doi-aqui" target="_blank">10.XXXX/tu-doi-aqui</a> <i>(Pending assignment)</i>'
    }
}

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="{{ lang }}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ t.title }}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background-color: #f4f7f6; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        .card { border-radius: 15px; box-shadow: 0 10px 20px rgba(0,0,0,0.1); border: none; }
        #loading { display: none; }
        .hero { background: linear-gradient(135deg, #0d6efd 0%, #0dcaf0 100%); color: white; border-radius: 15px 15px 0 0; padding: 20px; text-align: center; position: relative;}
        .tutorial-box { background-color: #e9f7fe; border-left: 5px solid #0dcaf0; padding: 15px; border-radius: 5px; font-size: 0.95rem;}
        .credits-box { background-color: #f8f9fa; border: 1px solid #dee2e6; padding: 15px; border-radius: 8px; font-size: 0.9rem; color: #495057; }
        .lang-switch { position: absolute; top: 15px; right: 20px; }
    </style>
</head>
<body>
<div class="container mt-4 mb-5">
    <div class="row justify-content-center">
        <div class="col-md-8">
            <div class="card">
                <div class="hero">
                    <a href="{{ t.lang_url }}" class="btn btn-sm btn-light fw-bold lang-switch">{{ t.lang_btn }}</a>
                    <h2 class="mb-0 mt-3">{{ t.hero_title }}</h2>
                    <p class="mt-2 mb-0">{{ t.hero_sub }}</p>
                </div>
                <div class="card-body p-4">
                    <div class="tutorial-box mb-4">
                        <h5 class="fw-bold text-primary">{{ t.tut_title }}</h5>
                        <p class="mb-2">{{ t.tut_p }}</p>
                        <ol class="mb-0">
                            <li>{{ t.tut_li1 | safe }}</li>
                            <li>{{ t.tut_li2 | safe }}</li>
                            <li>{{ t.tut_li3 | safe }}</li>
                        </ol>
                    </div>

                    <form id="searchForm" action="/procesar" method="POST" onsubmit="showLoading()">
                        <input type="hidden" name="ui_lang" value="{{ lang }}">
                        <div class="mb-4">
                            <label for="api_key" class="form-label fw-bold">{{ t.lbl_api }}</label>
                            <input type="password" class="form-control" id="api_key" name="api_key" required placeholder="{{ t.ph_api }}">
                        </div>
                        <hr class="mb-4">
                        <div class="mb-4">
                            <label for="tema" class="form-label fw-bold">{{ t.lbl_tema }}</label>
                            <input type="text" class="form-control form-control-lg" id="tema" name="tema" required placeholder="{{ t.ph_tema }}">
                        </div>
                        <div class="row mb-4">
                            <div class="col-md-6">
                                <label for="idioma" class="form-label fw-bold">{{ t.lbl_idioma }}</label>
                                <select class="form-select" id="idioma" name="idioma">
                                    <option value="es" {% if lang == 'es' %}selected{% endif %}>{{ t.opt_es }}</option>
                                    <option value="en" {% if lang == 'en' %}selected{% endif %}>{{ t.opt_en }}</option>
                                </select>
                            </div>
                            <div class="col-md-6 mt-3 mt-md-0">
                                <label for="limite" class="form-label fw-bold">{{ t.lbl_limite }}</label>
                                <input type="number" class="form-control" id="limite" name="limite" value="10" min="1" max="50">
                            </div>
                        </div>
                        <button type="submit" class="btn btn-primary w-100 py-3 fs-5 fw-bold" id="btnSubmit">
                            {{ t.btn_submit }}
                        </button>
                    </form>

                    <div id="loading" class="text-center mt-5 mb-3">
                        <div class="spinner-border text-primary" style="width: 4rem; height: 4rem;" role="status"></div>
                        <h4 class="mt-4 text-primary fw-bold">{{ t.loading_title }}</h4>
                        <p class="text-muted">{{ t.loading_text | safe }}</p>
                    </div>

                    <div class="credits-box mt-5">
                        <h6 class="fw-bold mb-2 text-dark">{{ t.cred_title }}</h6>
                        <p class="mb-1">{{ t.cred_author | safe }}</p>
                        <p class="mb-2"><i>{{ t.cred_cite }}</i></p>
                        <div class="p-2 bg-white border rounded">
                            {{ t.cred_doi | safe }}
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
</div>
<script>
    function showLoading() {
        document.getElementById('searchForm').style.display = 'none';
        document.getElementById('loading').style.display = 'block';
        setTimeout(() => {
            document.getElementById('searchForm').style.display = 'block';
            document.getElementById('loading').style.display = 'none';
        }, 30000); 
    }
</script>
</body>
</html>
"""

API_URL = "https://api.groq.com/openai/v1/chat/completions"
MODELO = "llama3-70b-8192" 

def llamar_ia(prompt, user_api_key, system_prompt="Eres un investigador experto."):
    payload = {"model": MODELO, "messages":[{"role": "system", "content": system_prompt}, {"role": "user", "content": prompt}], "temperature": 0.2}
    headers = {"Authorization": f"Bearer {user_api_key}", "Content-Type": "application/json"}
    try:
        r = requests.post(API_URL, headers=headers, json=payload, timeout=25)
        r.raise_for_status()
        return r.json()['choices'][0]['message']['content']
    except Exception as e:
        print(f"⚠️ Error API Groq: {e}")
        return ""

def generar_keywords(tema, user_api_key, idioma="es"):
    lang_str = "ESPAÑOL" if idioma == "es" else "INGLÉS"
    prompt = f"Genera 8 palabras clave científicas en {lang_str} para investigar sobre '{tema}'. Responde SOLO con las palabras separadas por comas."
    res = llamar_ia(prompt, user_api_key)
    return [k.strip() for k in res.split(",")] if res else [tema] 

def obtener_abstract_scholar(doi):
    try:
        url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}?fields=abstract"
        r = requests.get(url, timeout=5)
        if r.status_code == 200:
            res = r.json().get("abstract")
            return res if res else ""
    except: pass
    return ""

def extraer_de_web(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        r = requests.get(url, headers=headers, timeout=15, allow_redirects=True, verify=False)
        if r.status_code != 200: return "", False
        soup = BeautifulSoup(r.text, 'html.parser')
        meta_tags =['citation_abstract', 'dc.description', 'og:description', 'twitter:description', 'description']
        for tag in meta_tags:
            meta_abs = soup.find('meta', attrs={'name': re.compile(f'^{tag}$', re.I)}) or soup.find('meta', attrs={'property': re.compile(f'^{tag}$', re.I)})
            if meta_abs and meta_abs.get('content') and len(meta_abs.get('content')) > 60:
                return str(meta_abs['content']), True
        texto_limpio = soup.get_text(separator=' ', strip=True)
        match = re.search(r'(?i)\b(abstract|resumen|summary)\b[\s\S]{100,3500}', texto_limpio)
        if match: return match.group(0), True 
        parrafos = soup.find_all('p')
        texto_bruto = " ".join([p.get_text(strip=True) for p in parrafos if len(p.get_text(strip=True)) > 60])
        if len(texto_bruto) > 150: return texto_bruto[:3500], False 
    except Exception: pass
    return "", False

def extraer_de_pdf(pdf_url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        r = requests.get(pdf_url, headers=headers, timeout=15, verify=False, allow_redirects=True)
        if 'application/pdf' not in r.headers.get('Content-Type', '').lower() and not pdf_url.endswith('.pdf'): return "", False
        f = BytesIO(r.content)
        lector = PyPDF2.PdfReader(f)
        texto = ""
        for i in range(min(3, len(lector.pages))):
            pagina_texto = lector.pages[i].extract_text()
            if pagina_texto: texto += pagina_texto + " \n"
        texto_limpio = re.sub(r'\s+', ' ', texto).strip()
        match = re.search(r'(?i)\b(abstract|resumen|summary)\b[\s\S]{100,3500}', texto_limpio)
        if match: return match.group(0), True
        return texto_limpio[:3500], False
    except Exception: pass
    return "", False

def buscar_referencias_reales(keywords, limite=20): 
    query = "+".join(keywords).replace(' ', '+')
    url = "https://api.crossref.org/works"
    params = {"query": query, "rows": limite + 10, "filter": "type:journal-article", "sort": "relevance"}
    referencias =[]
    try:
        r = requests.get(url, params=params, timeout=30)
        items = r.json().get("message", {}).get("items",[])
        for idx, i in enumerate(items, 1):
            if len(referencias) >= limite: break
            autores_cr = i.get("author",[])
            autores_str = ", ".join([f"{a.get('given', '')} {a.get('family', '')}".strip() for a in autores_cr]) or "Anon"
            titulo = str(i.get("title",["Sin título"])[0]).replace("&", "y")
            revista = str(i.get("container-title",["Revista Académica"])[0])
            doi_raw = str(i.get("DOI", "Sin DOI"))
            doi_link = f"https://doi.org/{doi_raw}" if doi_raw != "Sin DOI" else "Sin DOI"
            
            enlaces = i.get("link",[])
            pdf_url = "No disponible"
            for link in enlaces:
                if link.get("content-type") == "application/pdf":
                    pdf_url = link.get("URL"); break
            
            texto_paper = ""
            fuente = ""
            texto_crudo = i.get("abstract", "")
            if texto_crudo:
                texto_paper = re.sub(r'<[^>]+>', '', str(texto_crudo)).strip()
                fuente = "API CrossRef"
            if len(texto_paper) < 50 and doi_raw != "Sin DOI":
                res = obtener_abstract_scholar(doi_raw)
                if res: texto_paper = str(res).strip(); fuente = "API Semantic Scholar"
            if len(texto_paper) < 50 and doi_link != "Sin DOI":
                texto, _ = extraer_de_web(doi_link)
                if len(texto) > 50: texto_paper = texto; fuente = "Web Scraping"
            if len(texto_paper) < 50 and pdf_url.startswith("http"):
                texto, _ = extraer_de_pdf(pdf_url)
                if len(texto) > 50: texto_paper = texto; fuente = "PDF Scraping"
            if len(texto_paper) < 50:
                texto_paper = "Texto no disponible por bloqueo de la revista."
                fuente = "Protegido/Pago"
                
            referencias.append({
                "autores": autores_str, "titulo": titulo, "revista": revista, 
                "doi": doi_link, "pdf": pdf_url, "texto_articulo": texto_paper, "fuente_extraccion": fuente
            })
    except Exception as e:
        print(f"Error en scraping: {e}")
    return referencias

def crear_excel_en_memoria(referencias, ui_lang='es'):
    wb = openpyxl.Workbook()
    ws = wb.active
    if ui_lang == 'en':
        ws.title = "Bibliography"
        encabezados =["Authors", "Article Title", "Journal", "DOI Link", "PDF Download", "Extraction Source", "Extracted Abstract / Summary"]
    else:
        ws.title = "Bibliografía"
        encabezados =["Autores", "Título del Artículo", "Revista", "Enlace DOI", "Descarga PDF", "Origen del Texto", "Abstract / Resumen Extraído"]
    ws.append(encabezados)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font = Font(bold=True)
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
    for ref in referencias:
        texto_crudo = ref['texto_articulo'][:4000] if ref['texto_articulo'] else "Texto no disponible."
        ws.append([ref['autores'], ref['titulo'], ref['revista'], ref['doi'], ref['pdf'], ref['fuente_extraccion'], texto_crudo])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 90
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

app = Flask(__name__)

@app.route('/')
def index():
    lang = request.args.get('lang', 'es')
    if lang not in ['es', 'en']: lang = 'es'
    return render_template_string(HTML_TEMPLATE, lang=lang, t=TRANSLATIONS[lang])

@app.route('/procesar', methods=['POST'])
def procesar():
    user_api_key = request.form.get('api_key').strip()
    tema = request.form.get('tema')
    idioma = request.form.get('idioma', 'es')
    ui_lang = request.form.get('ui_lang', 'es')
    try: limite = int(request.form.get('limite', 10))
    except ValueError: limite = 10
    
    kws = generar_keywords(tema, user_api_key, idioma)
    referencias = buscar_referencias_reales(kws, limite=limite) 
    excel_file = crear_excel_en_memoria(referencias, ui_lang)
    
    prefijo = "Research" if ui_lang == 'en' else "Investigacion"
    nombre_archivo = f"{prefijo}_{tema.replace(' ', '_')[:20]}.xlsx"
    return send_file(excel_file, download_name=nombre_archivo, as_attachment=True)

if __name__ == '__main__':
    # Para pruebas locales
    app.run(port=5000, use_reloader=False)
